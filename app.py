"""
Amazon Kindle Bestseller Scraper — Flask Backend
Run: python app.py  →  open http://localhost:5000
"""

import asyncio, io, json, os, queue, re, threading, time, uuid
from datetime import datetime
from flask import Flask, jsonify, request, send_file, Response, render_template
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

app = Flask(__name__)
CORS(app)

# In-memory job store: job_id -> { status, progress_queue, books, error }
JOBS: dict[str, dict] = {}


# ─── CLEANING HELPERS ─────────────────────────────────────────────────────────

def clean_price(raw):
    if not raw: return ""
    m = re.search(r"\$[\d,]+\.?\d*", raw)
    return m.group(0) if m else raw.strip()

def clean_rating(raw):
    if not raw: return ""
    m = re.search(r"([\d.]+)\s*out of", str(raw))
    if m: return float(m.group(1))
    m = re.search(r"([\d.]+)", str(raw))
    return float(m.group(1)) if m else ""

def clean_reviews(raw):
    if not raw: return ""
    text = str(raw).strip()

    # Handle formats like "1,234 ratings", "(2,567)", "3.2K ratings"
    # Extract number part

    # Remove commas and parentheses
    cleaned = re.sub(r'[,.()]+', '', text)

    # Handle K (thousand) notation
    if 'K' in cleaned.upper():
        num = re.search(r'(\d+(?:\.\d+)?)', cleaned)
        if num:
            return int(float(num.group(1)) * 1000)

    # Handle M (million) notation
    if 'M' in cleaned.upper():
        num = re.search(r'(\d+(?:\.\d+)?)', cleaned)
        if num:
            return int(float(num.group(1)) * 1000000)

    # Find any number in the text
    num_match = re.search(r'(\d+)', cleaned)
    if num_match:
        return int(num_match.group(1))

    return ""

def clean_date(raw):
    if not raw: return ""
    months = {"January":"01","February":"02","March":"03","April":"04",
               "May":"05","June":"06","July":"07","August":"08",
               "September":"09","October":"10","November":"11","December":"12"}
    m = re.match(r"(\w+)\s+(\d+),\s+(\d{4})", raw.strip())
    if m:
        mn, d, y = m.groups()
        return f"{y}-{months.get(mn,'01')}-{int(d):02d}"
    m = re.match(r"(\w+)\s+(\d{4})", raw.strip())
    if m:
        mn, y = m.groups()
        mo = months.get(mn, "")
        return f"{y}-{mo}" if mo else raw.strip()
    return raw.strip()

def clean_url(href):
    if not href: return ""
    full = f"https://www.amazon.com{href}" if href.startswith("/") else href
    asin = re.search(r"/dp/([A-Z0-9]{10})", full)
    return f"https://www.amazon.com/dp/{asin.group(1)}" if asin else full


# ─── SCRAPER (ASYNC) ──────────────────────────────────────────────────────────

async def scrape_listing_page(page, url, log):
    books = []
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=35000)
        await asyncio.sleep(2)
    except Exception as e:
        log(f"⚠️  Could not load listing page: {e}")
        return books

    if "captcha" in page.url.lower() or "ap/signin" in page.url.lower():
        log("⚠️  Amazon CAPTCHA detected. Try again or use a logged-in browser profile.")
        return books

    items = await page.query_selector_all(
        "li.zg-item-immersion, .p13n-sc-uncoverable-faceout"
    )
    log(f"📄  Found {len(items)} items on page")

    for idx, item in enumerate(items, 1):
        book = {"rank": idx}
        try:
            rb = await item.query_selector(".zg-bdg-text")
            if rb:
                rt = await rb.inner_text()
                book["rank"] = int(rt.replace("#","").strip())
        except: pass

        for sel in [
            "._cDEzb_p13n-sc-css-line-clamp-1_1Fn1y",
            "span._cDEzb_p13n-sc-css-line-clamp-2_EWgj1",
            ".p13n-sc-line-clamp-2",
            "a .p13n-sc-line-clamp-1",
        ]:
            try:
                el = await item.query_selector(sel)
                if el:
                    t = (await el.inner_text()).strip()
                    if t: book["title"] = t; break
            except: pass

        for sel in [".a-size-small.a-link-child", ".a-size-small a"]:
            try:
                el = await item.query_selector(sel)
                if el:
                    a = (await el.inner_text()).strip()
                    if a: book["author"] = a; break
            except: pass

        for sel in [
            ".a-icon-star-small .a-icon-alt",
            ".a-icon-alt[aria-label*='star']",
            "[aria-label*='stars']"
        ]:
            try:
                el = await item.query_selector(sel)
                if el:
                    raw = await el.get_attribute("aria-label") or await el.inner_text()
                    book["rating"] = clean_rating(raw)
                    break
            except: pass

        for sel in [
            "a[href*='customer-reviews']",
            ".a-link-normal[href*='reviews']",
            "span[data-hook='total-review-count']",
            ".a-declarative .a-link-normal[href*='reviews']",
            ".a-link-emphasis[href*='reviews']"
        ]:
            try:
                el = await item.query_selector(sel)
                if el:
                    raw_text = await el.inner_text()
                    log(f"  Found review element: '{raw_text}'")
                    book["num_reviews"] = clean_reviews(raw_text)
                    log(f"  Cleaned reviews: {book['num_reviews']}")
                    break
            except Exception as e:
                log(f"  Review extraction error: {e}")

        for sel in [
            "span._cDEzb_p13n-sc-price_3mJ9Z",
            ".a-color-price", "span.p13n-sc-price"
        ]:
            try:
                el = await item.query_selector(sel)
                if el:
                    p = (await el.inner_text()).strip()
                    if p: book["price"] = clean_price(p); break
            except: pass

        try:
            el = await item.query_selector("a.a-link-normal")
            if el:
                href = await el.get_attribute("href")
                if href: book["url"] = clean_url(href)
        except: pass

        if "title" in book or "url" in book:
            books.append(book)

    return books


async def scrape_book_page(page, book, log):
    url = book.get("url", "")
    if not url: return book
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(1.2)

        # Extract description with multiple fallback selectors
        for sel in [
            "#bookDescription_feature_div .a-expander-content",
            "#productDescription",
            "#bookDescription_feature_div span",
            "#bookDescription_feature_div",
            "#productDescriptionFeatureDiv",
            "#description-iframe-container",
            "div[data-feature-name='bookDescription']",
            "div#bookDescription_feature_div > div",
            "div#productDescription > div",
        ]:
            try:
                el = await page.query_selector(sel)
                if el:
                    t = (await el.inner_text()).strip()
                    if t and len(t) > 20:  # Ensure meaningful description
                        book["description"] = t[:900]
                        break
            except: pass

        # Extract publisher from product details section
        if "publisher" not in book or not book["publisher"]:
            for sel in [
                "#productDetails_feature_div",
                "#productDetailsTable",
                "#detailBullets_feature_div",
                "#productDetails",
                "#detail-bullets",
                ".a-section.a-spacing-medium.a-spacing-top-small",
                "div#productDetails_feature_div > div > table",
            ]:
                try:
                    el = await page.query_selector(sel)
                    if el:
                        details_text = await el.inner_text()
                        # Look for publisher in details section
                        pub_patterns = [
                            r"Publisher\s*[:\u200e]\s*([^\n\r]+)",
                            r"Publisher\s*[:]\s*([^\n\r]+)",
                            r"Publisher\s*([^\n\r]+)",
                        ]
                        for pat in pub_patterns:
                            m = re.search(pat, details_text, re.IGNORECASE)
                            if m:
                                publisher = m.group(1).strip()
                                # Remove common suffixes and special characters
                                publisher = re.sub(r'\s*\([^)]*\)$', '', publisher)
                                publisher = re.sub(r'\s*\[[^\]]*\]$', '', publisher)
                                publisher = re.sub(r'^[\u200e\u200f\s:]+', '', publisher)  # Remove RTL chars and leading spaces/colons
                                publisher = re.sub(r'^\s*[:]\s*', '', publisher)  # Remove leading colon
                                book["publisher"] = publisher[:80]
                                break
                        if "publisher" in book and book["publisher"]:
                            break
                except: pass

        # Extract publication date from product details section
        if "publication_date" not in book or not book["publication_date"]:
            for sel in [
                "#productDetails_feature_div",
                "#productDetailsTable",
                "#detailBullets_feature_div",
                "#productDetails",
                "#detail-bullets",
                ".a-section.a-spacing-medium.a-spacing-top-small",
                "div#productDetails_feature_div > div > table",
            ]:
                try:
                    el = await page.query_selector(sel)
                    if el:
                        details_text = await el.inner_text()
                        # Look for publication date in details section
                        date_patterns = [
                            r"Publication date\s*[:\u200e]\s*([^\n\r]+)",
                            r"Publication date\s*[:]\s*([^\n\r]+)",
                            r"Publication date\s*([^\n\r]+)",
                            r"Published on\s*[:\u200e]\s*([^\n\r]+)",
                            r"Published on\s*[:]\s*([^\n\r]+)",
                            r"Published on\s*([^\n\r]+)",
                            r"First published\s*[:]\s*([^\n\r]+)",
                        ]
                        for pat in date_patterns:
                            m = re.search(pat, details_text, re.IGNORECASE)
                            if m:
                                book["publication_date"] = clean_date(m.group(1).strip())
                                # Clean up special characters from date
                                book["publication_date"] = re.sub(r'^[\u200e\u200f\s:]+', '', book["publication_date"])
                                book["publication_date"] = re.sub(r'^\s*[:]\s*', '', book["publication_date"])
                                break
                        if "publication_date" in book and book["publication_date"]:
                            break
                except: pass
    except Exception as e:
        log(f"  ⚠️  {book.get('title','?')[:40]}: {e}")
    return book


async def run_scrape(job_id: str, url: str, scrape_details: bool, max_books: int):
    job = JOBS[job_id]
    log = lambda msg: job["q"].put({"type": "log", "msg": msg})

    try:
        from playwright.async_api import async_playwright
        log("🚀  Launching browser...")

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-blink-features=AutomationControlled"]
            )
            ctx = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1280, "height": 900},
                locale="en-US",
            )
            page = await ctx.new_page()
            await page.add_init_script(
                "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
            )

            # Paginate until we hit max_books (Amazon ZGBS: ~30 per page, up to 5 pages)
            books = []
            base_url = url.split("?")[0].rstrip("/")
            page_num = 1
            while len(books) < max_books:
                if page_num == 1:
                    page_url = base_url
                else:
                    page_url = f"{base_url}/ref=zg_bs_pg_{page_num}?_encoding=UTF8&pg={page_num}"
                log(f"📋  Scraping listing page {page_num}...")
                new_books = await scrape_listing_page(page, page_url, log)
                if not new_books:
                    log(f"  ⚠️  No more books found on page {page_num}, stopping")
                    break
                books += new_books
                log(f"  ✅  Page {page_num}: {len(new_books)} books (running total: {len(books)})")
                page_num += 1
                if page_num > 5:
                    break
                await asyncio.sleep(1.5)

            books = books[:max_books]
            log(f"✅  Collected {len(books)} book entries total")

            if scrape_details:
                log(f"🔍  Visiting individual book pages for details...")
                for i, book in enumerate(books):
                    title_short = book.get("title", "Unknown")[:45]
                    log(f"  [{i+1}/{len(books)}] {title_short}...")
                    books[i] = await scrape_book_page(page, book, log)
                    job["q"].put({"type": "progress",
                                  "val": int(50 + (i+1)/len(books)*50)})

            await browser.close()

        # Ensure all books have all fields
        FIELDS = ["rank","title","author","rating","num_reviews",
                  "price","url","description","publisher","publication_date"]
        for b in books:
            for f in FIELDS:
                b.setdefault(f, "")

        job["books"] = books
        job["status"] = "done"
        log(f"🎉  Scraping complete! {len(books)} books ready.")
        job["q"].put({"type": "done", "count": len(books)})

    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)
        log(f"❌  Fatal error: {e}")
        job["q"].put({"type": "error", "msg": str(e)})


# ─── EXCEL BUILDER ────────────────────────────────────────────────────────────

def build_excel(books: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bestseller Data"

    HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    ALT_FILL  = PatternFill("solid", fgColor="EBF0FA")
    WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")
    THIN = Side(style="thin", color="C8D0E0")
    BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    COLS = [
        ("Rank", 7), ("Title", 44), ("Author", 22), ("Rating", 9),
        ("# Reviews", 12), ("Price", 10), ("URL", 48),
        ("Description", 52), ("Publisher", 26), ("Publication Date", 17),
    ]
    KEYS = ["rank","title","author","rating","num_reviews","price",
            "url","description","publisher","publication_date"]

    for ci, (name, w) in enumerate(COLS, 1):
        c = ws.cell(1, ci, name)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = CTR; c.border = BDR
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    for ri, book in enumerate(books, 2):
        fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
        for ci, key in enumerate(KEYS, 1):
            c = ws.cell(ri, ci, book.get(key, ""))
            c.fill = fill; c.border = BDR
            c.font = Font(name="Calibri", size=10)
            c.alignment = CTR if ci in (1,4,5,6,10) else LFT
        ws.row_dimensions[ri].height = 52

    # Color scale on Rating
    if len(books) > 1:
        ws.conditional_formatting.add(
            f"D2:D{len(books)+1}",
            ColorScaleRule(start_type="min", start_color="FFF2CC",
                           end_type="max",  end_color="00B050")
        )

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 50

    def s2(r, c, v, bold=False, bg=None):
        cell = ws2.cell(r, c, v)
        cell.font = Font(name="Calibri", bold=bold, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        ws2.row_dimensions[r].height = 20
        return cell

    s2(1,1,"Amazon Kindle Bestsellers — Scraped Dataset Summary", bold=True, bg="1F3864").font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws2.merge_cells("A1:B1"); ws2.row_dimensions[1].height = 28

    rated = [b for b in books if isinstance(b.get("rating"), float)]
    reviewed = [b for b in books if isinstance(b.get("num_reviews"), int)]
    priced = [b for b in books if b.get("price","").startswith("$")]

    stats = [
        ("Total books scraped", len(books)),
        ("Books with rating", len(rated)),
        ("Average rating", round(sum(b["rating"] for b in rated)/len(rated),2) if rated else "N/A"),
        ("Books with review count", len(reviewed)),
        ("Total reviews (sum)", f"{sum(b['num_reviews'] for b in reviewed):,}" if reviewed else "N/A"),
        ("Books with price", len(priced)),
        ("Scrape timestamp", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for i, (k, v) in enumerate(stats, 3):
        s2(i, 1, k, bold=True, bg="D9E1F2")
        s2(i, 2, v)

    # Observations
    obs_start = len(stats) + 5
    for i, (title, text) in enumerate([
        ("Observation 1 — Pricing Tiers",
         "Titles cluster in 3 price bands: KU/Amazon Originals (~$0-2.99), indie ($3.99-5.99), and traditional publishers ($7.99-9.99). KU titles dominate top 5 due to velocity from subscriber reads."),
        ("Observation 2 — Publisher Concentration",
         "Top 5 often controlled by 1-2 publishers (Amazon Originals + 1 major). Below rank 10, fragmentation is high — indie/self-pub authors compete directly with Big-5."),
        ("Observation 3 — Rank vs Review Count",
         "BSR is velocity-weighted (last 24-72hr sales), not cumulative. New releases from high-follower authors instantly top charts despite far fewer reviews than legacy titles."),
    ], 0):
        r = obs_start + i * 3
        s2(r,   1, title, bold=True, bg="FFF2CC"); ws2.merge_cells(f"A{r}:B{r}")
        c = ws2.cell(r+1, 1, text)
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws2.merge_cells(f"A{r+1}:B{r+1}")
        ws2.row_dimensions[r+1].height = 55

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── ROUTES ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/favicon.ico")
def favicon():
    path = os.path.join(app.root_path, "static", "favicon.ico")
    if os.path.exists(path):
        return send_file(path, mimetype="image/x-icon")
    png_path = os.path.join(app.root_path, "static", "favicon.png")
    if os.path.exists(png_path):
        return send_file(png_path, mimetype="image/png")
    return ("", 204)


@app.route("/favicon.png")
def favicon_png():
    png_path = os.path.join(app.root_path, "static", "favicon.png")
    if os.path.exists(png_path):
        return send_file(png_path, mimetype="image/png")
    return ("", 404)


@app.route("/api/scrape", methods=["POST"])
def start_scrape():
    data = request.json or {}
    url = data.get("url", "").strip()
    if not url:
        return jsonify({"error": "URL is required"}), 400
    if "amazon.com" not in url:
        return jsonify({"error": "Please enter a valid Amazon URL"}), 400

    job_id = str(uuid.uuid4())
    JOBS[job_id] = {
        "status": "running",
        "q": queue.Queue(),
        "books": [],
        "error": None,
    }

    def thread_target():
        asyncio.run(run_scrape(
            job_id, url,
            scrape_details=data.get("scrape_details", True),
            max_books=int(data.get("max_books", 100)),
        ))

    t = threading.Thread(target=thread_target, daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


@app.route("/api/stream/<job_id>")
def stream(job_id):
    if job_id not in JOBS:
        return jsonify({"error": "Job not found"}), 404

    def generate():
        job = JOBS[job_id]
        while True:
            try:
                msg = job["q"].get(timeout=30)
                yield f"data: {json.dumps(msg)}\n\n"
                if msg["type"] in ("done", "error"):
                    break
            except queue.Empty:
                yield f"data: {json.dumps({'type':'ping'})}\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache",
                             "X-Accel-Buffering": "no"})


@app.route("/api/preview/<job_id>")
def preview(job_id):
    if job_id not in JOBS:
        return jsonify({"error": "Job not found"}), 404
    job = JOBS[job_id]
    all_books = job["books"]
    return jsonify({
        "status": job["status"],
        "count": len(all_books),
        "stats": {
            "with_price":       sum(1 for b in all_books if b.get("price","")),
            "with_rating":      sum(1 for b in all_books if b.get("rating") not in ("", None)),
            "with_description": sum(1 for b in all_books if b.get("description","").strip()),
            "with_publisher":   sum(1 for b in all_books if b.get("publisher","").strip()),
        },
        "books": all_books[:5],
    })


@app.route("/api/download/<job_id>")
def download(job_id):
    if job_id not in JOBS:
        return jsonify({"error": "Job not found"}), 404
    job = JOBS[job_id]
    if not job["books"]:
        return jsonify({"error": "No data to download"}), 400

    buf = build_excel(job["books"])
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"kindle_bestsellers_{ts}.xlsx"
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/download_csv/<job_id>")
def download_csv(job_id):
    if job_id not in JOBS:
        return jsonify({"error": "Job not found"}), 404
    job = JOBS[job_id]
    if not job["books"]:
        return jsonify({"error": "No data to download"}), 400

    import csv
    KEYS = ["rank","title","author","rating","num_reviews","price",
            "url","description","publisher","publication_date"]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=KEYS, extrasaction="ignore")
    w.writeheader()
    w.writerows(job["books"])
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=kindle_bestsellers_{ts}.csv"}
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    is_local = port == 5000
    if is_local:
        print("\n🚀  Amazon Kindle Bestseller Scraper")
        print(f"   Open http://localhost:{port} in your browser\n")
    app.run(debug=False, host="0.0.0.0", port=port, threaded=True)
